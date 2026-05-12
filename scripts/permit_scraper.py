#!/usr/bin/env python3
"""
Maricopa County Weekly Permit Activity Report scraper.

Usage:

    # Discover available reports and print to stdout
    python scripts/permit_scraper.py --discover

    # Discover and save index CSV
    python scripts/permit_scraper.py --discover --output-dir=data/permit-activity

    # Download reports (uses archive_index.csv to decide what to fetch)
    python scripts/permit_scraper.py --download
    python scripts/permit_scraper.py --download --limit 3
    python scripts/permit_scraper.py --download --start-date=2026-01-01 --end-date=2026-06-01
    python scripts/permit_scraper.py --download --force   # re-download even if present

    # Inspect the 3 newest downloaded reports
    python scripts/permit_scraper.py --inspect

    # All in one: discover, download (limit 5), inspect
    python scripts/permit_scraper.py --discover --download --limit 5 --inspect

    # Parse downloaded reports into the database
    python scripts/permit_scraper.py --init-db
    python scripts/permit_scraper.py --sync --limit 3

    # Summary reports
    python scripts/permit_scraper.py --summary
    python scripts/permit_scraper.py --summary --by month
    python scripts/permit_scraper.py --summary --by city
    python scripts/permit_scraper.py --summary --by contractor
"""

import argparse
import csv
import datetime
import hashlib
import io
import os
import re
import sys
import time
import urllib.request
from pathlib import Path
from typing import Optional

# lazy-import db so this module can be loaded without the full stack
def _get_db():
    import db
    return db


# Lazy imports for spreadsheet reading
def _get_openpyxl():
    import openpyxl
    return openpyxl


def _get_xlrd():
    import xlrd
    return xlrd
import urllib.error


# ── Constants ───────────────────────────────────────────────────────────────

ARCHIVE_URL = "https://www.maricopa.gov/Archive.aspx?AMID=128"
BASE_URL = "https://www.maricopa.gov"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)
DEFAULT_OUTPUT_DIR = "data/permit-activity"
INDEX_FILENAME = "archive_index.csv"
RAW_SUBDIR = "raw"


# ── Date parsing ────────────────────────────────────────────────────────────

# Month name → number mapping
MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

# Ordinal suffix pattern (1st, 2nd, 3rd, 4th, ...) — we strip these before parsing
ORDINAL_RE = re.compile(r"(\d+)(?:st|nd|rd|th)", re.I)


def _parse_report_date(text: str) -> Optional[str]:
    """Parse a date string like 'May 4, 2026' or 'December 19th, 2022'
    and return YYYY-MM-DD.

    Returns None if the date cannot be parsed.
    """
    text = text.strip()
    # Strip ordinal suffixes: 19th → 19
    text = ORDINAL_RE.sub(r"\1", text)
    # Remove stray commas around spaces: "November27,, 2023" → "November 27, 2023"
    text = re.sub(r"(\d)\s*,?\s*,?\s*(\d{4})", r"\1, \2", text)
    text = re.sub(r",\s*,", ",", text)
    # Clean internal spaces: "October 16 , 2023" → "October 16, 2023"
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+,", ",", text)

    # Try: "Month Day, Year"
    m = re.match(
        r"([A-Za-z]+)\s+(\d{1,2}),?\s*(\d{4})",
        text,
    )
    if m:
        month_name = m.group(1).lower()
        day = int(m.group(2))
        year = int(m.group(3))
        month = MONTH_MAP.get(month_name)
        if month and 1 <= day <= 31 and 2000 <= year <= 2099:
            return f"{year:04d}-{month:02d}-{day:02d}"

    return None


def _extract_links_from_archive(html: str) -> list[dict]:
    """Parse the archive page HTML and extract report links.

    Returns a list of dicts with keys: report_title, report_date, archive_url, adid.
    Dates are returned as YYYY-MM-DD strings.
    """
    records: list[dict] = []
    seen_adids: set[str] = set()

    # Structure:
    #   <a href="Archive.aspx?ADID=XXXX">
    #     <span>Weekly Permit Activity Report May 4, 2026  </span>
    #   </a>
    # Titles may include "(XLS)" suffix.
    link_pattern = re.compile(
        r'<a\s+href="Archive\.aspx\?ADID=(\d+)"[^>]*>'
        r'\s*<span[^>]*>(Weekly\s+Permit\s+Activity\s+Report\b.*?)</span>',
        re.I,
    )

    for adid, title_raw in link_pattern.findall(html):
        title_raw = title_raw.strip()

        if adid in seen_adids:
            continue
        seen_adids.add(adid)

        # Remove "(XLS)" marker before date extraction
        title_clean = re.sub(r"\s*\(XLS\)\s*", "", title_raw, flags=re.I).strip()
        title_for_date = title_clean  # use the cleaned title for date parsing

        # Extract date from the trailing portion of the title.
        # Handle odd spacing: "November27,, 2023" (no space), "October 16 , 2023" (space before comma)
        date_str = None
        date_match = re.search(
            r"([A-Za-z]+)[,\s]*(\d{1,2})(?:st|nd|rd|th)?[,\s]*(\d{4})\s*$",
            title_for_date,
        )
        if date_match:
            combined = f"{date_match.group(1)} {date_match.group(2)}, {date_match.group(3)}"
            date_str = _parse_report_date(combined)
        if not date_str:
            date_str = _parse_report_date(title_for_date)

        records.append({
            "report_title": title_clean,
            "report_date": date_str or "",
            "archive_url": f"{BASE_URL}/Archive.aspx?ADID={adid}",
            "adid": adid,
        })

    # Sort by date descending (newest first)
    records.sort(key=lambda r: (r["report_date"] or "0000-00-00", r["adid"]), reverse=True)
    return records


# ── File-type detection ─────────────────────────────────────────────────────

def _detect_file_type(content: bytes, content_type: Optional[str] = None) -> str:
    """Detect file type from magic bytes (preferred) or Content-Type header."""
    # Magic bytes take priority — they tell us what the content actually is.
    if content[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return "xls"
    if content[:4] == b"\x50\x4B\x03\x04":
        return "xlsx"
    if content[:5] == b"\x25\x50\x44\x46\x2D":
        return "pdf"

    # Fallback to Content-Type header
    if content_type:
        ct = content_type.lower()
        if "spreadsheet" in ct and "openxml" in ct:
            return "xlsx"
        if "vnd.ms-excel" in ct or "xls" in ct:
            return "xls"
        if "pdf" in ct:
            return "pdf"
        if "html" in ct:
            return "html"
    if content[:4] == b"\xEF\xBB\xBF" or content[:2] in (b"\xFF\xFE", b"\xFE\xFF"):
        return "csv"

    return "unknown"


# ── HTML fetch helpers ──────────────────────────────────────────────────────

def _fetch_html(url: str) -> str:
    """Fetch a URL and return its text content."""
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8", errors="replace")


def _fetch_raw(url: str) -> tuple[bytes, Optional[str]]:
    """Fetch a URL and return (raw_bytes, content_type)."""
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        ct = resp.headers.get("Content-Type")
        return resp.read(), ct


def _resolve_download_url(archive_url: str) -> tuple[str, bytes, Optional[str]]:
    """Given an archive URL (Archive.aspx?ADID=XXXX), follow the redirect
    to the ViewFile URL and return (viewfile_url, content_bytes, content_type).

    The ViewFile URL serves the actual file (XLSX / XLS / PDF).
    """
    req = urllib.request.Request(archive_url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            content = resp.read()
            ct = resp.headers.get("Content-Type")
            final_url = resp.url
            return final_url, content, ct
    except urllib.error.HTTPError as e:
        print(f"  HTTP {e.code} fetching {archive_url}", file=sys.stderr)
        return archive_url, b"", None


# ── CSV index I/O ───────────────────────────────────────────────────────────

def _index_path(output_dir: str) -> Path:
    return Path(output_dir) / INDEX_FILENAME


def _load_index(output_dir: str) -> list[dict]:
    """Load existing archive_index.csv, returning list of dicts."""
    path = _index_path(output_dir)
    if not path.exists():
        return []
    with open(path, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _save_index(records: list[dict], output_dir: str):
    """Write records to archive_index.csv."""
    path = _index_path(output_dir)
    path.parent.mkdir(parents=True, exist_ok=True)

    now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
    for r in records:
        if "scraped_at" not in r:
            r["scraped_at"] = now_iso

    fieldnames = [
        "report_title", "report_date", "archive_url",
        "resolved_download_url", "adid",
        "file_name", "file_type",
        "scraped_at",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        # Deduplicate by adid, keep newest first
        seen: set[str] = set()
        for r in records:
            adid = r.get("adid", "")
            if adid and adid not in seen:
                seen.add(adid)
                writer.writerow(r)


# ── Download ────────────────────────────────────────────────────────────────

def _download_record(
    record: dict,
    output_dir: str,
    force: bool = False,
) -> dict:
    """Download a single report, save to disk, and return updated record with
    resolved_download_url, file_name, and file_type.

    Idempotent: skips if the file already exists and force is False.
    """
    adid = record.get("adid", "")
    date_str = record.get("report_date", "")

    # Resolve download URL and get content
    archive_url = record["archive_url"]
    final_url, content, content_type = _resolve_download_url(archive_url)
    record["resolved_download_url"] = final_url

    # Determine file extension
    file_type = _detect_file_type(content, content_type)
    record["file_type"] = file_type

    ext_map = {"xlsx": ".xlsx", "xls": ".xls", "pdf": ".pdf", "csv": ".csv"}
    ext = ext_map.get(file_type, ".bin")

    # Generate filename: YYYY-MM-DD_ADID.ext
    date_prefix = date_str if date_str else "unknown_date"
    file_name = f"{date_prefix}_{adid}{ext}"
    record["file_name"] = file_name

    # Determine save path: raw/YYYY/YYYY-MM-DD/
    year = date_str[:4] if len(date_str) >= 4 else "unknown"
    save_dir = Path(output_dir) / RAW_SUBDIR / year / date_str
    save_path = save_dir / file_name

    # Check if file already exists and has the same content
    if save_path.exists() and not force:
        existing_hash = hashlib.sha256(save_path.read_bytes()).hexdigest()
        new_hash = hashlib.sha256(content).hexdigest()
        if existing_hash == new_hash:
            print(f"  Skipped {file_name} (already exists, unchanged)")
            return record

    # Save
    save_dir.mkdir(parents=True, exist_ok=True)
    save_path.write_bytes(content)
    print(f"  Downloaded {file_name} ({file_type}, {len(content):,} bytes)")
    return record


def _filter_records_by_date(
    records: list[dict],
    start_date: Optional[str],
    end_date: Optional[str],
) -> list[dict]:
    """Filter records by ISO date range (inclusive)."""
    if not start_date and not end_date:
        return records
    filtered = []
    for r in records:
        rd = r.get("report_date", "")
        if not rd:
            continue
        if start_date and rd < start_date:
            continue
        if end_date and rd > end_date:
            continue
        filtered.append(r)
    return filtered


# ── Inspect ─────────────────────────────────────────────────────────────────

def _inspect_report(path: Path):
    """Print file type and apparent column headers from a report."""
    content = path.read_bytes()
    file_type = _detect_file_type(content)
    print(f"\n  File: {path.name}")
    print(f"  Type: {file_type}")
    print(f"  Size: {len(content):,} bytes")

    if file_type == "xlsx":
        # Attempt to read first sheet via zipfile + xml; fallback to strings
        try:
            import zipfile
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                # Find the shared strings and first sheet
                sheet_names = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
                if sheet_names:
                    sheet_xml = z.read(sheet_names[0]).decode("utf-8", errors="replace")
                    # Extract text between <c>...</c> tags as a crude header extraction
                    header_cells = re.findall(r'<c[^>]*>.*?<v>(.*?)</v>', sheet_xml[:5000])
                    if header_cells:
                        print(f"  First cell values: {header_cells[:10]}")
                    # Try to read shared strings for more readable output
                    try:
                        ss_xml = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
                        strings = re.findall(r"<si>.*?<t>(.*?)</t>.*?</si>", ss_xml, re.DOTALL)
                        if strings and header_cells:
                            resolved = [strings[int(s)] if s.isdigit() and int(s) < len(strings) else s for s in header_cells]
                            print(f"  Headers (resolved): {resolved[:15]}")
                    except KeyError:
                        pass
                    # Print raw XML snippet of first 3 rows
                    rows = re.findall(r"<row[^>]*>(.*?)</row>", sheet_xml[:10000], re.DOTALL)
                    print(f"  Rows in first sheet (approx): {len(rows)}")
        except ImportError:
            print("  (install 'openpyxl' for detailed XLSX inspection)")
        except Exception as e:
            print(f"  XLSX parse error: {e}")
    elif file_type == "xls":
        print("  (install 'xlrd' for XLS inspection)")
    elif file_type == "pdf":
        print("  (install 'pdfplumber' or 'tabula' for PDF inspection)")
    else:
        # Attempt to print first few lines for text-based formats
        try:
            text = content.decode("utf-8", errors="replace")
            lines = text.strip().split("\n")[:5]
            for i, line in enumerate(lines):
                print(f"  Line {i}: {line[:200]}")
        except Exception:
            pass


# ── XLSX Parser ─────────────────────────────────────────────────────────────

# Known column name patterns for header detection (case-insensitive)
_HEADER_PATTERNS = {
    "permit_type": ["permit type"],
    "work_class": ["work class", "permit work class"],
    "permit_number": ["permit number", "permit #", "permit no", "tracking number"],
    "permit_issue_date": ["permit issue date", "issue date", "permit date", "date issued"],
    "permit_description": ["permit description", "description", "work description"],
    "permit_valuation": ["permit valuation", "valuation", "estimated cost", "value"],
    "permit_square_feet": ["permit square feet", "square feet", "sq ft", "sqft", "square footage", "building footprint square footage"],
    "parcel_no": ["parcel no", "parcel number", "parcel #", "apn", "assessor parcel"],
    "no_units": ["no units", "units", "number of units"],
    "job_address": ["job address", "address", "site address", "location"],
    "subdivision": ["subdivision", "sub"],
    "lot": ["lot", "lot number"],
    "job_city": ["job city", "city"],
    "job_state": ["job state", "state"],
    "job_zip": ["job zip", "zip", "zip code"],
    "owner_name": ["owner name", "owner", "property owner"],
    "contractor_name": ["contractor name", "contractor", "builder"],
    "contractor_phone": ["contractor phone", "phone", "contractor phone number", "contactor phone number"],
    "contractor_email": ["contractor email", "email", "contractor email address", "contact email"],
    "contractor_phone": ["contractor phone", "phone", "contractor phone number", "contactor phone number", "contact business phone"],

    # 2024+ columns not caught above
    "application_date": ["permit application date"],
    "job_zip": ["job zip", "zip", "zip code", "postal code"],
}


# Old-format (2012-2023) header overrides for columns whose names collide
# with modern patterns. Detected by the presence of 'THIS PERMIT IS FOR'.
_OLD_FORMAT_PATTERNS = {
    "native_category": ["description"],
    "work_class": ["category"],
    "permit_description": ["this permit is for"],
}


def _clean_val(val):
    """Normalize a cell value to a clean string or None."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.isoformat()[:10]
    if isinstance(val, float):
        s = str(val)
        if s.endswith(".0"):
            s = s[:-2]
        return s
    s = str(val).strip()
    if not s or s == "None":
        return None
    return s


def _detect_header_row(rows: list) -> int:
    """Find the row index containing column headers.

    Scans for a row that matches at least 3 known column-name patterns.
    Falls back to row 1 if nothing is found.
    """
    all_patterns = set()
    for pats in _HEADER_PATTERNS.values():
        all_patterns.update(pats)

    best_idx = 1  # fallback
    best_score = 0
    for ri, row in enumerate(rows):
        if not row:
            continue
        score = 0
        for cell in row:
            if cell is None:
                continue
            low = str(cell).strip().lower()
            for pat in all_patterns:
                if low == pat:
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_idx = ri
        if score >= 5:  # good enough, stop early
            break
    return best_idx


def _is_old_format(headers: list) -> bool:
    """Detect pre-2024 permit files by checking for legacy column headers."""
    low = [str(c).strip().lower() for c in headers if c is not None]
    return "this permit is for" in low or "tracking number" in low or "building footprint square footage" in low


def _build_column_map(headers: list) -> dict:
    """Map column name strings to position index (0-based).

    Handles inconsistent column layouts by matching known patterns.
    Unmatched columns are mapped to their position with a generic key.
    """
    col_map: dict[str, int] = {}
    for ci, cell in enumerate(headers):
        if cell is None:
            continue
        low = str(cell).strip().lower()
        matched = False
        for field_name, patterns in _HEADER_PATTERNS.items():
            if low in patterns:
                col_map[field_name] = ci
                matched = True
                break
        if not matched:
            col_map[f"_col_{ci}"] = ci

    # Apply old-format overrides — the 'DESCRIPTION' header means 'native_category'
    # in old files, not 'permit_description'.
    if _is_old_format(headers):
        for field_name, patterns in _OLD_FORMAT_PATTERNS.items():
            for ci, cell in enumerate(headers):
                if cell is None:
                    continue
                low = str(cell).strip().lower()
                if low in patterns:
                    col_map[field_name] = ci
                    # Remove the standard-field key that collided
                    for std_field in _HEADER_PATTERNS:
                        if col_map.get(std_field) == ci and std_field != field_name:
                            del col_map[std_field]
                            break

    return col_map


def _compute_row_hash(report_date: str, row: dict) -> str:
    """Generate a stable SHA-256 row hash for dedup.

    Based on: report_date + parcel_no + job_address + permit_description
    + valuation + square_feet + owner_name + contractor_name.
    """
    parts = [
        report_date or "",
        row.get("parcel_no") or "",
        row.get("job_address") or "",
        row.get("permit_description") or "",
        str(row.get("permit_valuation") or ""),
        str(row.get("permit_square_feet") or ""),
        row.get("owner_name") or "",
        row.get("contractor_name") or "",
    ]
    raw = "||".join(parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _read_rows_xlsx(filepath: str) -> list[list]:
    """Read all rows from an XLSX file using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _read_rows_xls(filepath: str) -> list[list]:
    """Read all rows from a legacy XLS file using xlrd."""
    import os
    import xlrd
    wb = xlrd.open_workbook(filepath, logfile=open(os.devnull, "w"))
    ws = wb.sheet_by_index(0)
    rows = []
    for ri in range(ws.nrows):
        row = []
        for ci in range(ws.ncols):
            cell = ws.cell(ri, ci)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_tuple(cell.value, wb.datemode))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def parse_spreadsheet(filepath: str) -> tuple[list[dict], list[str]]:
    """Parse an XLSX or XLS file into structured permit rows.

    Automatically detects format by file extension.
    Returns (rows, warnings) where:
    - rows: list of dicts with field_name keys
    - warnings: list of human-readable warning strings about parsing
    """
    from pathlib import Path

    warnings: list[str] = []
    ext = Path(filepath).suffix.lower()

    # Convert xlrd date tuples to ISO strings (same format openpyxl returns)
    def _clean_xls_val(v):
        if isinstance(v, tuple) and len(v) >= 3:
            # xlrd date tuple: (year, month, day, hour, minute, second)
            try:
                return datetime.date(v[0], v[1], v[2]).isoformat()
            except (ValueError, IndexError):
                return str(v)
        return v

    try:
        if ext == ".xls":
            all_rows = _read_rows_xls(filepath)
            # Convert date tuples to ISO strings
            all_rows = [[_clean_xls_val(v) for v in row] for row in all_rows]
            warnings.append("Parsed as legacy .xls format")
        else:
            all_rows = _read_rows_xlsx(filepath)
    except Exception as e:
        return [], [f"Failed to read spreadsheet: {e}"]

    if not all_rows:
        return [], ["Empty spreadsheet"]

    # Detect header row
    header_idx = _detect_header_row(all_rows)
    headers = list(all_rows[header_idx]) if header_idx < len(all_rows) else []
    if header_idx > 0:
        warnings.append(f"Header detected at row {header_idx} (skipped {header_idx} title/pre-header rows)")

    col_map = _build_column_map(headers)
    fields_found = [k for k in col_map if not k.startswith("_")]
    if not fields_found:
        return [], ["No known column headers found in row {}".format(header_idx)]
    warnings.append(f"Resolved columns: {', '.join(sorted(fields_found))}")

    # Extract report title from the first row if it looks like one
    report_title = ""
    if all_rows[0] and all_rows[0][0]:
        first_cell = str(all_rows[0][0]).strip()
        if "weekly" in first_cell.lower() or "permit" in first_cell.lower():
            report_title = first_cell

    # Parse data rows (skip header row and any rows before it)
    parsed: list[dict] = []
    for ri in range(header_idx + 1, len(all_rows)):
        row = all_rows[ri]
        if not row or all(c is None for c in row):
            continue

        record: dict[str, str] = {}
        for field_name, ci in col_map.items():
            if field_name.startswith("_"):
                continue
            val = row[ci] if ci < len(row) else None
            record[field_name] = _clean_val(val)

        parsed.append(record)

    return parsed, warnings


def _file_content_hash(filepath: str) -> Optional[str]:
    """SHA-256 of file contents."""
    try:
        return hashlib.sha256(Path(filepath).read_bytes()).hexdigest()
    except Exception:
        return None


# ── DB helpers ──────────────────────────────────────────────────────────────

def _init_db():
    """Create/update permit tables."""
    from db import PermitReport, Permit
    from db import init_db
    init_db()
    print("Database initialized (permit_reports, permits tables ready)", file=sys.stderr)


def _normalize_permit_category(permit_type: Optional[str],
                                  native_category: Optional[str] = None) -> str:
    """Map permit_type or native_category to a cross-jurisdiction category.

    New-format files (2024+) have permit_type ("Building (Residential)").
    Old-format files (2012-2023) have native_category ("Residential", "Fence").
    """
    src = permit_type or native_category
    if not src:
        return "Other"
    src_lower = src.strip().lower()
    if "residential" in src_lower:
        return "Residential"
    if "commercial" in src_lower:
        return "Commercial"
    if "industrial" in src_lower:
        return "Industrial"
    return "Other"


def _sync_single_report(session, record: dict, output_dir: str) -> int:
    """Parse one downloaded XLSX, store rows in DB.

    Idempotent: deletes existing permits for this report_adid first.
    Returns count of rows inserted.
    """
    db_mod = _get_db()
    PermitReport = db_mod.PermitReport
    Permit = db_mod.Permit

    adid = record.get("adid", "")
    file_name = record.get("file_name", "")
    file_type = record.get("file_type", "")
    report_date = record.get("report_date", "")
    source_url = record.get("archive_url", "")

    # Locate the file on disk
    if file_name:
        year = report_date[:4] if len(report_date) >= 4 else "unknown"
        local_path = Path(output_dir) / RAW_SUBDIR / year / report_date / file_name
    else:
        return 0

    if not local_path.exists():
        print(f"  File not found: {local_path}", file=sys.stderr)
        return 0

    # Parse
    rows, warnings = parse_spreadsheet(str(local_path))
    if not rows:
        for w in warnings:
            print(f"  Warning: {w}", file=sys.stderr)
        return 0

    # Upsert permit report record
    content_hash = _file_content_hash(str(local_path))
    now = datetime.datetime.now(datetime.timezone.utc)

    from sqlalchemy import select
    existing_report = session.execute(
        select(PermitReport).where(PermitReport.adid == adid)
    ).scalar_one_or_none()

    if existing_report:
        existing_report.file_type = file_type
        existing_report.file_name = file_name
        existing_report.local_path = str(local_path)
        existing_report.content_hash = content_hash
        existing_report.downloaded_at = now
        existing_report.row_count = len(rows)
        existing_report.updated_at = now
    else:
        session.add(PermitReport(
            report_date=report_date,
            adid=adid,
            report_title=record.get("report_title", ""),
            file_type=file_type,
            file_name=file_name,
            source_url=source_url,
            local_path=str(local_path),
            content_hash=content_hash,
            downloaded_at=now,
            row_count=len(rows),
        ))

    # Delete existing permits for this report (idempotent replace)
    session.execute(
        Permit.__table__.delete().where(Permit.report_adid == adid)
    )
    session.flush()  # ensure DELETE runs before ORM INSERTs

    # Insert new permits
    inserted = 0
    seen_hashes: set[str] = set()
    for row in rows:
        # Compute row_hash (always, for metadata and future dedup)
        permit_number = row.get("permit_number") or None
        row_hash = _compute_row_hash(report_date, row)

        # Skip duplicate rows within this report (same adid + same content)
        if row_hash in seen_hashes:
            continue
        seen_hashes.add(row_hash)

        # Determine uniqueness key
        if permit_number:
            uniq_key = permit_number
        else:
            uniq_key = row_hash

        # Build permit_number lookup: check if already exists
        # (delete/replace semantics: for this report_adid, we already deleted
        # old rows. But other reports might have the same permit_number.
        # This shouldn't happen for weekly reports — each permit appears once.)

        permit_type = row.get("permit_type")

        rec = Permit(
            report_date=report_date,
            report_adid=adid,
            source_file=file_name,
            permit_type=permit_type,
            work_class=row.get("work_class"),
            permit_number=permit_number,
            permit_issue_date=row.get("permit_issue_date"),
            permit_description=row.get("permit_description"),
            permit_valuation=row.get("permit_valuation"),
            permit_square_feet=row.get("permit_square_feet"),
            parcel_no=row.get("parcel_no"),
            no_units=row.get("no_units"),
            job_address=row.get("job_address"),
            subdivision=row.get("subdivision"),
            lot=row.get("lot"),
            job_city=row.get("job_city"),
            job_state=row.get("job_state"),
            job_zip=row.get("job_zip"),
            owner_name=row.get("owner_name"),
            contractor_name=row.get("contractor_name"),
            contractor_phone=row.get("contractor_phone"),
            contractor_email=row.get("contractor_email"),
            row_hash=row_hash,
            jurisdiction="Maricopa County",
            normalized_category=_normalize_permit_category(permit_type, row.get("native_category")),
            native_type=permit_type,
            native_category=row.get("native_category"),
        )
        session.add(rec)
        inserted += 1

    session.commit()
    return inserted


def _parse_num(val: Optional[str]) -> float:
    """Parse a numeric string, returning 0 on failure.

    Handles commas, leading dollar signs, whitespace.
    """
    if not val:
        return 0.0
    s = str(val).replace("$", "").replace(",", "").replace(" ", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _print_summary(session, group_by: Optional[str] = None):
    """Print a summary of all synced permits."""
    db_mod = _get_db()
    Permit = db_mod.Permit
    PermitReport = db_mod.PermitReport
    from sqlalchemy import select, func, text

    total = session.execute(select(func.count()).select_from(Permit)).scalar() or 0
    if total == 0:
        print("No permits in database. Run --sync first.")
        return

    # Column names
    report_count = session.execute(
        select(func.count()).select_from(PermitReport)
    ).scalar() or 0

    print(f"\n{'='*72}")
    print(f"  PERMIT SUMMARY")
    print(f"  {total} permits across {report_count} weekly reports")
    print(f"{'='*72}")

    # Overall stats
    rows = session.execute(
        select(Permit.permit_valuation, Permit.permit_square_feet, Permit.no_units)
    ).all()

    total_val = 0.0
    total_sqft = 0.0
    total_units = 0
    for r in rows:
        total_val += _parse_num(r.permit_valuation)
        total_sqft += _parse_num(r.permit_square_feet)
        total_units += int(float(r.no_units)) if r.no_units and r.no_units.strip() else 0

    print(f"  {'Permit count:':30s} {total:,}")
    print(f"  {'Total valuation:':30s} ${total_val:,.0f}")
    print(f"  {'Total sq ft:':30s} {total_sqft:,.0f}")
    print(f"  {'Total units:':30s} {total_units:,}")
    print(f"  {'Avg valuation/sqft:':30s} ${total_val / total_sqft:.2f}" if total_sqft > 0 else f"  {'Avg valuation/sqft:':30s} N/A")
    print()

    # Group-by options
    if group_by == "month":
        rows = session.execute(
            text("""
                SELECT SUBSTR(report_date, 1, 7) as month,
                       COUNT(*) as cnt,
                       COALESCE(SUM(CAST(REPLACE(permit_valuation, ',', '') AS REAL)), 0) as val,
                       COALESCE(SUM(CAST(REPLACE(permit_square_feet, ',', '') AS REAL)), 0) as sqft,
                       COALESCE(SUM(CAST(COALESCE(NULLIF(no_units, ''), '0') AS REAL)), 0) as units
                FROM permits
                GROUP BY month
                ORDER BY month DESC
                LIMIT 24
            """)
        ).all()
        print(f"  {'Month':12s} {'# Permits':>10s} {'Valuation':>14s} {'Sq Ft':>12s} {'Units':>8s}")
        print(f"  {'-'*56}")
        for r in rows:
            print(f"  {r.month:12s} {r.cnt:>10,} ${r.val:>11,.0f} {r.sqft:>11,.0f} {r.units:>8,.0f}")

    elif group_by == "city":
        rows = session.execute(
            text("""
                SELECT UPPER(COALESCE(NULLIF(job_city, ''), '(unknown)')) as city,
                       COUNT(*) as cnt,
                       COALESCE(SUM(CAST(REPLACE(permit_valuation, ',', '') AS REAL)), 0) as val,
                       COALESCE(SUM(CAST(REPLACE(permit_square_feet, ',', '') AS REAL)), 0) as sqft
                FROM permits
                GROUP BY city
                ORDER BY sqft DESC
                LIMIT 20
            """)
        ).all()
        print(f"  {'City':20s} {'# Permits':>10s} {'Valuation':>14s} {'Sq Ft':>12s}")
        print(f"  {'-'*56}")
        for r in rows:
            print(f"  {r.city:20s} {r.cnt:>10,} ${r.val:>11,.0f} {r.sqft:>11,.0f}")

    elif group_by == "type":
        rows = session.execute(
            text("""
                SELECT UPPER(COALESCE(NULLIF(permit_type, ''), '(unknown)')) as ptype,
                       COUNT(*) as cnt,
                       COALESCE(SUM(CAST(REPLACE(permit_valuation, ',', '') AS REAL)), 0) as val,
                       COALESCE(SUM(CAST(REPLACE(permit_square_feet, ',', '') AS REAL)), 0) as sqft,
                       COALESCE(SUM(CAST(COALESCE(NULLIF(no_units, ''), '0') AS REAL)), 0) as units
                FROM permits
                GROUP BY ptype
                ORDER BY sqft DESC
                LIMIT 20
            """)
        ).all()
        print(f"  {'Permit Type':35s} {'#':>5s} {'Valuation':>14s} {'Sq Ft':>12s} {'Units':>8s}")
        print(f"  {'-'*74}")
        for r in rows:
            pt = r.ptype[:33] + ".." if len(r.ptype) > 33 else r.ptype
            print(f"  {pt:35s} {r.cnt:>5} ${r.val:>11,.0f} {r.sqft:>11,.0f} {r.units:>8,.0f}")
        print()

        # Also show cross-tab: top cities for each major type
        for major_type in ["BUILDING (RESIDENTIAL)", "BUILDING (COMMERCIAL)", "SIGN", "DEMOLITION"]:
            sub = session.execute(
                text("""
                    SELECT UPPER(COALESCE(NULLIF(job_city, ''), '(unknown)')) as city,
                           COUNT(*) as cnt,
                           COALESCE(SUM(CAST(REPLACE(permit_square_feet, ',', '') AS REAL)), 0) as sqft
                    FROM permits
                    WHERE UPPER(COALESCE(NULLIF(permit_type, ''), '(unknown)')) LIKE :pattern
                    GROUP BY city
                    ORDER BY sqft DESC
                    LIMIT 5
                """),
                {"pattern": f"%{major_type}%"},
            ).all()
            print(f"  Top cities for {major_type}:")
            print(f"    {'City':25s} {'#':>5s} {'Sq Ft':>12s}")
            print(f"    {'-'*42}")
            for r in sub:
                print(f"    {r.city:25s} {r.cnt:>5} {r.sqft:>11,.0f}")
            print()

    elif group_by == "contractor":
        rows = session.execute(
            text("""
                SELECT COALESCE(NULLIF(contractor_name, ''), '(unknown)') as contractor,
                       COUNT(*) as cnt,
                       COALESCE(SUM(CAST(REPLACE(permit_valuation, ',', '') AS REAL)), 0) as val,
                       COALESCE(SUM(CAST(REPLACE(permit_square_feet, ',', '') AS REAL)), 0) as sqft
                FROM permits
                GROUP BY contractor
                ORDER BY sqft DESC
                LIMIT 20
            """)
        ).all()
        print(f"  {'Contractor':30s} {'# Permits':>10s} {'Valuation':>14s} {'Sq Ft':>12s}")
        print(f"  {'-'*66}")
        for r in rows:
            cname = r.contractor[:28] + ".." if len(r.contractor) > 28 else r.contractor
            print(f"  {cname:30s} {r.cnt:>10,} ${r.val:>11,.0f} {r.sqft:>11,.0f}")

    else:
        # Top permit descriptions by sq ft
        rows = session.execute(
            text("""
                SELECT COALESCE(NULLIF(permit_description, ''), '(blank)') as descr,
                       COUNT(*) as cnt,
                       COALESCE(SUM(CAST(REPLACE(permit_square_feet, ',', '') AS REAL)), 0) as sqft
                FROM permits
                GROUP BY descr
                ORDER BY sqft DESC
                LIMIT 10
            """)
        ).all()
        print(f"  {'Permit Description':50s} {'#':>5s} {'Sq Ft':>12s}")
        print(f"  {'-'*67}")
        for r in rows:
            d = r.descr[:48] + ".." if len(r.descr) > 48 else r.descr
            print(f"  {d:50s} {r.cnt:>5} {r.sqft:>11,.0f}")

    print()


# ── CLI ─────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Maricopa County Weekly Permit Activity Report scraper",
    )
    parser.add_argument(
        "--discover", action="store_true",
        help="Enumerate all report links from the archive page",
    )
    parser.add_argument(
        "--download", action="store_true",
        help="Download reports (uses archive_index.csv)",
    )
    parser.add_argument(
        "--inspect", action="store_true",
        help="Print file type and column headers from the newest 3 downloaded reports",
    )
    parser.add_argument(
        "--init-db", action="store_true",
        help="Create/update permit database tables",
    )
    parser.add_argument(
        "--sync", action="store_true",
        help="Parse downloaded reports into the database",
    )
    parser.add_argument(
        "--reparse", action="store_true",
        help="Re-parse all existing local files into the database (shortcut for --sync with no date/limit filter)",
    )
    parser.add_argument(
        "--summary", action="store_true",
        help="Print permit summary",
    )
    parser.add_argument(
        "--by", choices=["month", "city", "contractor", "type"], default=None,
        help="Group summary output by this dimension (used with --summary)",
    )
    parser.add_argument(
        "--output-dir", default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory (default: {DEFAULT_OUTPUT_DIR})",
    )
    parser.add_argument(
        "--limit", type=int, default=None,
        help="Maximum number of reports to process",
    )
    parser.add_argument(
        "--start-date", default=None,
        help="Earliest report date (inclusive, YYYY-MM-DD)",
    )
    parser.add_argument(
        "--end-date", default=None,
        help="Latest report date (inclusive, YYYY-MM-DD)",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Re-download even if files already exist",
    )
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    output_dir = args.output_dir
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # ── Init DB ─────────────────────────────────────────────────────────
    if args.init_db:
        _init_db()
        return

    # ── Discover ────────────────────────────────────────────────────────
    if args.discover:
        print(f"Fetching archive page: {ARCHIVE_URL}", file=sys.stderr)
        html = _fetch_html(ARCHIVE_URL)
        records = _extract_links_from_archive(html)
        print(f"Found {len(records)} reports", file=sys.stderr)

        # Filter by date range
        if args.start_date or args.end_date:
            records = _filter_records_by_date(records, args.start_date, args.end_date)
            print(f"After date filter: {len(records)} reports", file=sys.stderr)

        # Apply limit (before saving — latest records)
        if args.limit is not None:
            records = records[: args.limit]
            print(f"After limit: {len(records)} reports", file=sys.stderr)

        # Merge with existing index (preserve resolved URLs from prior runs)
        existing = _load_index(output_dir)
        existing_by_adid = {r["adid"]: r for r in existing if r.get("adid")}

        merged = []
        for r in records:
            adid = r.get("adid", "")
            if adid and adid in existing_by_adid:
                old = existing_by_adid[adid]
                for key in ("resolved_download_url", "file_name", "file_type", "scraped_at"):
                    if old.get(key):
                        r[key] = old[key]
            merged.append(r)

        _save_index(merged, output_dir)
        print(f"Saved index to {_index_path(output_dir)}", file=sys.stderr)

        # Print to stdout
        writer = csv.DictWriter(
            sys.stdout,
            fieldnames=["report_date", "adid", "report_title", "file_type", "file_name"],
            extrasaction="ignore",
        )
        writer.writeheader()
        for r in merged:
            writer.writerow(r)

    # ── Download ────────────────────────────────────────────────────────
    if args.download:
        records = _load_index(output_dir)
        # Auto-discover if the index is empty or very small (unlikely to
        # represent the full archive).  The user can always pre-populate
        # with an explicit --discover for full control.
        if len(records) < 100:
            print("Index empty or incomplete — auto-discovering...", file=sys.stderr)
            html = _fetch_html(ARCHIVE_URL)
            fresh = _extract_links_from_archive(html)
            if fresh:
                # Preserve any resolved download URLs from the old index
                old_by_adid = {r["adid"]: r for r in records if r.get("adid")}
                for r in fresh:
                    adid = r.get("adid", "")
                    if adid in old_by_adid:
                        old = old_by_adid[adid]
                        for k in ("resolved_download_url", "file_name", "file_type"):
                            if old.get(k):
                                r[k] = old[k]
                records = fresh
                _save_index(records, output_dir)
                print(f"Discovered {len(records)} reports", file=sys.stderr)
            else:
                print("Auto-discover found nothing.", file=sys.stderr)

        # Filter by date range
        if args.start_date or args.end_date:
            records = _filter_records_by_date(records, args.start_date, args.end_date)

        # Apply limit (from the NEWEST records since they're sorted by date desc)
        if args.limit is not None:
            records = records[: args.limit]

        if args.limit and not (args.start_date or args.end_date):
            print(f"Downloading up to {args.limit} reports (newest first)", file=sys.stderr)
        else:
            print(f"Downloading {len(records)} reports", file=sys.stderr)

        updated = []
        for record in records:
            updated.append(_download_record(record, output_dir, force=args.force))
            time.sleep(2)  # rate-limit politeness

        # Update index with resolved URLs and file info
        _save_index(updated, output_dir)

    # ── Sync / Reparse ─────────────────────────────────────────────────
    if args.sync or args.reparse:
        records = _load_index(output_dir)
        if not records:
            print("No index found. Run --discover or --download first.", file=sys.stderr)
            sys.exit(1)

        db_mod = _get_db()
        db_mod.init_db()
        session = db_mod.get_session()

        # Filter by date range
        if args.start_date or args.end_date:
            records = _filter_records_by_date(records, args.start_date, args.end_date)

        # Apply limit
        if args.limit is not None:
            records = records[: args.limit]

        print(f"Syncing {len(records)} reports into database...", file=sys.stderr)
        total_rows = 0
        synced = 0
        for record in records:
            adid = record.get("adid", "")
            date_str = record.get("report_date", "")
            print(f"  Parsing {date_str}_{adid}...", file=sys.stderr)
            try:
                count = _sync_single_report(session, record, output_dir)
                if count > 0:
                    total_rows += count
                    synced += 1
                    print(f"    -> {count} permit rows", file=sys.stderr)
                else:
                    print(f"    -> 0 rows (file not found or empty)", file=sys.stderr)
            except Exception as e:
                print(f"    -> ERROR: {e}", file=sys.stderr)
                try:
                    session.rollback()
                except Exception:
                    pass
                # Re-create a fresh session for the next report
                session.close()
                db_mod = _get_db()
                db_mod.init_db()
                session = db_mod.get_session()

        session.close()
        print(f"\nDone: {synced} reports synced, {total_rows} total permit rows", file=sys.stderr)

    # ── Summary ───────────────────────────────────────────────────────────
    if args.summary:
        db_mod = _get_db()
        db_mod.init_db()
        session = db_mod.get_session()
        _print_summary(session, group_by=args.by)
        session.close()

    # ── Inspect ─────────────────────────────────────────────────────────
    if args.inspect:
        raw_dir = Path(output_dir) / RAW_SUBDIR
        if not raw_dir.exists():
            print("No raw downloads found. Run --download first.", file=sys.stderr)
            sys.exit(1)

        # Find all downloaded files, sorted by date-modified or path (newest first)
        all_files = sorted(
            raw_dir.rglob("*"),
            key=lambda p: p.stat().st_mtime if p.is_file() else 0,
            reverse=True,
        )
        files = [f for f in all_files if f.is_file() and f.suffix in (".xlsx", ".xls", ".pdf", ".csv")]

        if not files:
            print("No report files found under {raw_dir}.", file=sys.stderr)
            sys.exit(1)

        limit = min(3, len(files))
        print(f"Inspecting {limit} newest downloaded report(s):", file=sys.stderr)
        for f in files[:limit]:
            _inspect_report(f)
        print()


if __name__ == "__main__":
    main()
